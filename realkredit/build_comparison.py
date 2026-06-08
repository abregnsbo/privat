"""Build a Google-Sheets-importable comparison of the Jyske and Danske
mortgage offers (realkredit/jyske_bank_laan.pdf and danske_bank_laan.pdf).

Reproduces the yearly payment tables from each offer, lists the annual
'bidrag' (admin margin) separately, and computes the present value of the
bidrag stream discounted at a configurable rate (default 2.5% p.a.).

Output: realkredit/mortgage_comparison.xlsx (multiple sheets, live formulas).

Notes on data sources
---------------------
* Jyske offer prints a 'Bidrag/tillæg' column per year, so those values are
  taken verbatim. The first three calendar years (2026-2028) are listed
  quarterly in the offer and are aggregated here into calendar-year rows.
* Danske offer's yearly table does NOT break out bidrag, only ydelse and
  restgæld. The bidrag stream is therefore reconstructed from a quarterly
  amortisation of the loan (principal 2,834,000, 4.00% coupon, 0.2060%
  bidragssats, 120 quarterly payments). The reconstruction reproduces the
  offer's published 'Restgæld ultimo' to within ~1,000 kr (<0.04%) and the
  yearly ydelse to within ~10 kr, so the derived bidrag is reliable.
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

DISCOUNT_DEFAULT = 0.025

# ---------------------------------------------------------------------------
# Danske bidrag reconstruction (quarterly), validated against the offer table.
# ---------------------------------------------------------------------------
def danske_bidrag_by_year():
    P0 = 2834000.0
    i = 0.04 / 4
    b = 0.002060 / 4
    n = 120
    PMT = P0 * i / (1 - (1 + i) ** -n)
    bal = P0
    pay_years = [2026] * 2 + [y for y in range(2027, 2056) for _ in range(4)] + [2056] * 2
    bidrag = {}
    for k, py in enumerate(pay_years):
        renter = round(bal * i, 2)
        bid = round(bal * b, 2)
        afdrag = bal if k == n - 1 else round(PMT - renter, 2)
        bal = round(bal - afdrag, 2)
        bidrag[py] = bidrag.get(py, 0.0) + bid
    return {y: round(v) for y, v in bidrag.items()}

# ---------------------------------------------------------------------------
# Published yearly tables.  Each row: (year, ydelse_for_skat, ydelse_efter_skat,
# afdrag, renter, bidrag, restgaeld_ultimo).  None where derived/not published.
# ---------------------------------------------------------------------------

# Danske: published ydelse (før/efter skat) and restgæld ultimo (Nyt lån column,
# offer page 8). bidrag reconstructed; afdrag/renter derived below.
DANSKE_PUB = {
    2026: (84700, 69763, 2809096), 2027: (168397, 139082, 2758051),
    2028: (168290, 139520, 2704934), 2029: (168179, 139975, 2649660),
    2030: (168063, 140449, 2592142), 2031: (167943, 140942, 2532289),
    2032: (167818, 141455, 2470005), 2033: (167687, 141989, 2405192),
    2034: (167552, 142545, 2337747), 2035: (167411, 143123, 2267564),
    2036: (167264, 143725, 2194532), 2037: (167111, 144351, 2118533),
    2038: (166953, 145003, 2039449), 2039: (166787, 145681, 1957154),
    2040: (166615, 146387, 1871518), 2041: (166436, 147121, 1782404),
    2042: (166250, 147885, 1689672), 2043: (166056, 148680, 1593174),
    2044: (165854, 149507, 1492758), 2045: (165644, 150368, 1388265),
    2046: (165425, 151264, 1279530), 2047: (165198, 152197, 1166379),
    2048: (164961, 153167, 1048633), 2049: (164715, 154176, 926107),
    2050: (164459, 155227, 798606), 2051: (164192, 156320, 665927),
    2052: (163915, 157458, 527862), 2053: (163626, 158642, 384190),
    2054: (163326, 159873, 234685), 2055: (163013, 161155, 79109),
    2056: (80353, 80043, 0),
}

def build_danske_rows():
    bidrag = danske_bidrag_by_year()
    rows = []
    prev_rest = 2834000  # principal
    for y in sorted(DANSKE_PUB):
        yf, ye, rest = DANSKE_PUB[y]
        bid = bidrag[y]
        afdrag = prev_rest - rest
        renter = yf - afdrag - bid          # derived so the row reconciles
        rows.append((y, yf, ye, afdrag, renter, bid, rest))
        prev_rest = rest
    return rows

# Jyske 'med afdrag' (offer page 7). 2026-2028 aggregated from quarterly rows.
JYSKE_AFDRAG = [
    (2026, 89149, 72953, 26131, 59662, 3356, 2822869),
    (2027, 169808, 139364, 51348, 112152, 6309, 2771521),
    (2028, 169690, 139812, 53433, 110066, 6192, 2718087),
    (2029, 169569, 140280, 55603, 107896, 6069, 2662484),
    (2030, 169442, 140765, 57861, 105639, 5942, 2604623),
    (2031, 169310, 141271, 60210, 103289, 5810, 2544413),
    (2032, 169172, 141797, 62655, 100844, 5673, 2481757),
    (2033, 169029, 142345, 65199, 98300, 5529, 2416558),
    (2034, 168880, 142914, 67846, 95653, 5380, 2348712),
    (2035, 168725, 143507, 70601, 92898, 5226, 2278111),
    (2036, 168564, 144124, 73468, 90032, 5064, 2204643),
    (2037, 168396, 144766, 76451, 87048, 4896, 2128191),
    (2038, 168221, 145434, 79555, 83944, 4722, 2048636),
    (2039, 168040, 146129, 82786, 80714, 4540, 1965850),
    (2040, 167851, 146853, 86147, 77353, 4351, 1879703),
    (2041, 167654, 147606, 89645, 73855, 4154, 1790058),
    (2042, 167449, 148389, 93285, 70215, 3950, 1696774),
    (2043, 167236, 149204, 97073, 66427, 3737, 1599701),
    (2044, 167014, 150052, 101014, 62485, 3515, 1498687),
    (2045, 166784, 150935, 105116, 58384, 3284, 1393571),
    (2046, 166544, 151854, 109384, 54116, 3044, 1284187),
    (2047, 166294, 152809, 113825, 49674, 2794, 1170362),
    (2048, 166034, 153804, 118447, 45053, 2534, 1051915),
    (2049, 165763, 154839, 123256, 40243, 2264, 928658),
    (2050, 165482, 155916, 128261, 35238, 1982, 800397),
    (2051, 165189, 157037, 133469, 30030, 1689, 666928),
    (2052, 164884, 158203, 138888, 24611, 1384, 528040),
    (2053, 164567, 159417, 144528, 18972, 1067, 383512),
    (2054, 164400, 160801, 150396, 13103, 900, 233115),
    (2055, 164400, 162370, 156503, 6997, 900, 76612),
    (2056, 78193, 77787, 76612, 1131, 450, 0),
]

# Jyske 'afdragsfrihed 10 år' (offer page 14). 2026-2028 aggregated.
JYSKE_AF10 = [
    (2026, 63826, 47423, 0, 60427, 3399, 2879000),
    (2027, 121636, 90376, 0, 115160, 6476, 2879000),
    (2028, 121636, 90376, 0, 115160, 6476, 2879000),
    (2029, 121638, 90377, 0, 115160, 6478, 2879000),
    (2030, 121638, 90377, 0, 115160, 6478, 2879000),
    (2031, 121638, 90377, 0, 115160, 6478, 2879000),
    (2032, 121638, 90377, 0, 115160, 6478, 2879000),
    (2033, 121638, 90377, 0, 115160, 6478, 2879000),
    (2034, 121638, 90377, 0, 115160, 6478, 2879000),
    (2035, 121638, 90377, 0, 115160, 6478, 2879000),
    (2036, 168949, 137752, 47561, 114923, 6464, 2831439),
    (2037, 216098, 185749, 98009, 111800, 6289, 2733430),
    (2038, 215874, 186605, 101989, 107820, 6065, 2631441),
    (2039, 215641, 187496, 106130, 103679, 5832, 2525311),
    (2040, 215398, 188424, 110439, 99370, 5590, 2414872),
    (2041, 215146, 189389, 114924, 94885, 5337, 2299948),
    (2042, 214884, 190393, 119590, 90219, 5075, 2180358),
    (2043, 214611, 191438, 124446, 85363, 4802, 2055913),
    (2044, 214326, 192526, 129499, 80310, 4517, 1926414),
    (2045, 214031, 193657, 134757, 75052, 4222, 1791657),
    (2046, 213723, 194835, 140229, 69580, 3914, 1651428),
    (2047, 213402, 196060, 145922, 63886, 3594, 1505506),
    (2048, 213069, 197335, 151847, 57961, 3260, 1353659),
    (2049, 212722, 198662, 158013, 51796, 2914, 1195646),
    (2050, 212361, 200043, 164429, 45380, 2553, 1031217),
    (2051, 211986, 201480, 171105, 38703, 2177, 860111),
    (2052, 211595, 202975, 178053, 31756, 1786, 682058),
    (2053, 211188, 204531, 185283, 24526, 1380, 496775),
    (2054, 210792, 206169, 192806, 17003, 983, 303969),
    (2055, 210709, 208120, 200635, 9174, 900, 103335),
    (2056, 105337, 104823, 103335, 1553, 450, 0),
]

# Jyske 'afdragsfrihed 30 år' (offer page 21). 2026-2028 aggregated.
JYSKE_AF30 = [
    (2026, 67683, 50289, 0, 61183, 6501, 2915000),
    (2027, 128988, 95840, 0, 116600, 12388, 2915000),
    (2028, 128988, 95840, 0, 116600, 12388, 2915000),
] + [
    (y, 128989, 95839, 0, 116600, 12389, 2915000) for y in range(2029, 2056)
] + [
    (2056, 2979494, 2962919, 2915000, 58300, 6194, 0),
]

# ---------------------------------------------------------------------------
# Loan metadata for the summary sheet.
# ---------------------------------------------------------------------------
LOANS = [
    # key, sheet, label, hovedstol, rente%, bidragssats%, type, n_rows
    ("danske", "Danske", "Danske / Realkredit Danmark – Obligationslån 4% (med afdrag)",
     2834000, 4.00, 0.2060, "30 år, med afdrag"),
    ("jyske_afdrag", "Jyske_afdrag", "Jyske Realkredit – Fastrente 4% (med afdrag)",
     2849000, 4.00, 0.225, "30 år, med afdrag"),
    ("jyske_af10", "Jyske_AF10", "Jyske Realkredit – Fastrente 4% (afdragsfrihed 10 år)",
     2879000, 4.00, 0.225, "30 år, 10 års afdragsfrihed"),
    ("jyske_af30", "Jyske_AF30", "Jyske Realkredit – Fastrente 4% (afdragsfrihed 30 år)",
     2915000, 4.00, 0.425, "30 år, 30 års afdragsfrihed"),
]
ROWS = {
    "danske": build_danske_rows(),
    "jyske_afdrag": JYSKE_AFDRAG,
    "jyske_af10": JYSKE_AF10,
    "jyske_af30": JYSKE_AF30,
}

# ---------------------------------------------------------------------------
# Workbook construction
# ---------------------------------------------------------------------------
HEAD_FILL = PatternFill("solid", fgColor="1F4E78")
HEAD_FONT = Font(bold=True, color="FFFFFF")
SUB_FILL = PatternFill("solid", fgColor="D9E1F2")
BID_FILL = PatternFill("solid", fgColor="FFF2CC")
TOT_FILL = PatternFill("solid", fgColor="E2EFDA")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MONEY = "#,##0"
DEC = "0.0000"

COLS = ["År", "Ydelse før skat", "Ydelse efter skat", "Afdrag",
        "Renter", "Bidrag", "Restgæld ultimo",
        "Diskonteringsfaktor", "NV af bidrag"]

def style_header(cell):
    cell.fill = HEAD_FILL
    cell.font = HEAD_FONT
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
    cell.border = BORDER

def write_detail(ws, key, label):
    rows = ROWS[key]
    ws["A1"] = label
    ws["A1"].font = Font(bold=True, size=12)
    ws["A2"] = ("Bidrag = bidragssats anvendt på restgælden. "
                "Diskonteringsrente og diskontering: se fanen 'Summary'. "
                "Diskonteringsfaktor = 1/(1+rente)^(år−2026).")
    ws["A2"].font = Font(italic=True, size=9, color="808080")

    hr = 4
    for c, name in enumerate(COLS, start=1):
        cell = ws.cell(row=hr, column=c, value=name)
        style_header(cell)

    first = hr + 1
    for r, row in enumerate(rows, start=first):
        year, yf, ye, af, re, bid, rest = row
        ws.cell(row=r, column=1, value=year).border = BORDER
        for c, val in zip(range(2, 8), (yf, ye, af, re, bid, rest)):
            cell = ws.cell(row=r, column=c, value=val)
            cell.number_format = MONEY
            cell.border = BORDER
        # discount factor (col 8) references the rate cell on Summary
        df = ws.cell(row=r, column=8,
                     value=f"=1/(1+Summary!$B$2)^(A{r}-2026)")
        df.number_format = DEC
        df.border = BORDER
        # PV of bidrag (col 9) = bidrag * discount factor
        pv = ws.cell(row=r, column=9, value=f"=F{r}*H{r}")
        pv.number_format = MONEY
        pv.border = BORDER
        pv.fill = BID_FILL
        ws.cell(row=r, column=6).fill = BID_FILL

    last = first + len(rows) - 1
    tr = last + 1
    ws.cell(row=tr, column=1, value="Total").font = Font(bold=True)
    for c in (2, 3, 4, 5, 6):
        cl = get_column_letter(c)
        cell = ws.cell(row=tr, column=c, value=f"=SUM({cl}{first}:{cl}{last})")
        cell.number_format = MONEY
        cell.font = Font(bold=True)
        cell.fill = TOT_FILL
        cell.border = BORDER
    cell = ws.cell(row=tr, column=9, value=f"=SUM(I{first}:I{last})")
    cell.number_format = MONEY
    cell.font = Font(bold=True)
    cell.fill = TOT_FILL
    cell.border = BORDER
    for c in (1, 7, 8):
        ws.cell(row=tr, column=c).fill = TOT_FILL

    widths = [8, 15, 16, 12, 12, 12, 15, 16, 14]
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = ws.cell(row=first, column=2)
    return first, last, tr

def main():
    wb = openpyxl.Workbook()
    summary = wb.active
    summary.title = "Summary"

    # discount rate input cell
    summary["A1"] = "Sammenligning af realkreditlån – Jyske vs. Danske"
    summary["A1"].font = Font(bold=True, size=14)
    summary["A2"] = "Diskonteringsrente (p.a.)"
    summary["A2"].font = Font(bold=True)
    summary["B2"] = DISCOUNT_DEFAULT
    summary["B2"].number_format = "0.0%"
    summary["B2"].fill = PatternFill("solid", fgColor="FFF2CC")
    summary["B2"].font = Font(bold=True)
    summary["B2"].border = BORDER
    summary["C2"] = "← ændr her; alle nutidsværdier opdateres"
    summary["C2"].font = Font(italic=True, color="808080")

    # detail sheets first (so we know their row ranges)
    sheet_ranges = {}
    for key, sheet, label, *_ in LOANS:
        ws = wb.create_sheet(sheet)
        sheet_ranges[key] = (sheet, *write_detail(ws, key, label))

    # comparison table on Summary
    hdr = ["Lån", "Type", "Hovedstol", "Rente", "Bidragssats",
           "Bidrag nominelt (sum)", "NV af bidrag @ rente",
           "Samlet tilbagebetaling (før skat)"]
    r0 = 4
    for c, name in enumerate(hdr, start=1):
        cell = summary.cell(row=r0, column=c, value=name)
        style_header(cell)

    for idx, (key, sheet, label, hov, rente, bids, typ) in enumerate(LOANS):
        r = r0 + 1 + idx
        sname, first, last, tr = sheet_ranges[key]
        summary.cell(row=r, column=1, value=label).border = BORDER
        summary.cell(row=r, column=2, value=typ).border = BORDER
        c3 = summary.cell(row=r, column=3, value=hov); c3.number_format = MONEY; c3.border = BORDER
        c4 = summary.cell(row=r, column=4, value=rente / 100); c4.number_format = "0.00%"; c4.border = BORDER
        c5 = summary.cell(row=r, column=5, value=bids / 100); c5.number_format = "0.0000%"; c5.border = BORDER
        c6 = summary.cell(row=r, column=6, value=f"='{sname}'!F{tr}"); c6.number_format = MONEY; c6.border = BORDER
        c7 = summary.cell(row=r, column=7, value=f"='{sname}'!I{tr}")
        c7.number_format = MONEY; c7.border = BORDER; c7.fill = BID_FILL; c7.font = Font(bold=True)
        c8 = summary.cell(row=r, column=8, value=f"='{sname}'!B{tr}"); c8.number_format = MONEY; c8.border = BORDER

    # headline difference: Jyske med afdrag vs Danske (the comparable pair)
    rdiff = r0 + 1 + len(LOANS) + 1
    summary.cell(row=rdiff, column=1,
                 value="Forskel i NV af bidrag (Jyske med afdrag − Danske)").font = Font(bold=True)
    da_r = r0 + 1  # danske is first
    jy_r = r0 + 2  # jyske_afdrag is second
    d = summary.cell(row=rdiff, column=7, value=f"=G{jy_r}-G{da_r}")
    d.number_format = "#,##0;[Red]-#,##0"; d.font = Font(bold=True); d.fill = TOT_FILL

    notes = [
        "",
        "Noter:",
        "• Begge tilbud er 30-årige fastforrentede obligationslån med 4,00% kuponrente; "
        "den direkte sammenlignelige variant er 'med afdrag'.",
        "• Jyske-tilbuddet oplyser bidrag pr. år direkte i ydelsesforløbet – brugt som anført.",
        "• Danske-tilbuddet oplyser ikke bidrag i årstabellen; bidraget er rekonstrueret "
        "fra et kvartalsvist afdragsforløb (0,2060% p.a. af restgælden) og reproducerer "
        "tilbuddets 'Restgæld ultimo' inden for ~1.000 kr.",
        "• 2026-2028 vises i tilbuddene kvartalsvist og er her aggregeret til kalenderår.",
        "• Nutidsværdi (NV) diskonterer hvert års bidrag med 1/(1+rente)^(år−2026); "
        "2026 diskonteres ikke (t=0).",
        "• Afdragsfrihed holder restgælden høj længere og giver derfor markant højere "
        "samlet bidrag.",
    ]
    nr = rdiff + 2
    for i, line in enumerate(notes):
        cell = summary.cell(row=nr + i, column=1, value=line)
        if line.startswith("Noter"):
            cell.font = Font(bold=True)
        else:
            cell.font = Font(size=9, color="595959")

    widths = [62, 28, 13, 9, 13, 20, 20, 26]
    for c, w in enumerate(widths, start=1):
        summary.column_dimensions[get_column_letter(c)].width = w

    out = "mortgage_comparison.xlsx"
    wb.save(out)
    print("wrote", out)

if __name__ == "__main__":
    main()
